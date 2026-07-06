"""
Transformación del Reporte de Incidencias Qinaya
=================================================
Este script:
1. Lee la hoja original 'Reporte de Incidencias.xlsx'
2. Separa cada evento/incidente de la columna "Bot" en su propia fila
3. Extrae la fecha del texto a una columna independiente
4. Clasifica el tipo de evento (Incidencia Técnica, Capacitación, Seguimiento, etc.)
5. Agrega el código DANE de cada colegio
6. Genera un nuevo archivo 'Reporte_Incidencias_Normalizado.xlsx' listo para Google Sheets
"""

import openpyxl
import re
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# CÓDIGOS DANE POR COLEGIO
# ==========================================
CODIGOS_DANE = {
    "Colegio Eduardo Santos": "111001014001",
    "Colegio Santa Lucia": "111001098965",
    "Colegio Rodrigo Lara Bonilla": "111001036781",
    "Colegio Marco Tulio": "111001001121",
    "Colegio Manuela Beltran": "111001010731",
    "Colegio José Manuel Restrepo": "111001045535",
    "Colegio Ext Camilo Torres": "111001010839",
    "Colegio Atanasio Girardot": "111001012602",
    "Colegio Antonio García": "111001107859",
    "Colegio Paraíso Mirador": "111001047678",
    "Colegio Rural Pasquilla": "211850001317",
    "Colegio Liceo Nacional  Agustín Nieto Caballero": "111001025313",
    "Colegio El Salitre Suba": "111769000247",
    "Colegio Ciudad de Villavicencio": "211001094832",
    "Colegio Los Comuneros": "111001044270",
    "Colegio San francisco de Asis": "111001102181",
    "Colegio Sorrento": "111001020168",
    "Colegio Gustavo Restrepo": "111001027332",
    "Colegio Manuel Cepeda": "111001027391",
    "Colegio CEDID San Pablo (IED)": "111001015694",
    "Colegio Nuevo Horizonte": "111001086681",
    "Colegio Costa Rica Sede B": "111279000362",
    "Colegio Virginia Gutierrez de Pineda (IED)": "111001107069",
    "Colegio La joya": "111001098973",
    "Colegio Moralba Suroriental Sede A": "111001014389",
}


def buscar_dane(nombre_colegio):
    """Busca el código DANE por nombre exacto o aproximado."""
    if nombre_colegio in CODIGOS_DANE:
        return CODIGOS_DANE[nombre_colegio]
    # Búsqueda aproximada
    nombre_lower = nombre_colegio.lower()
    for key, val in CODIGOS_DANE.items():
        if key.lower() in nombre_lower or nombre_lower in key.lower():
            return val
    return "POR VERIFICAR"


def extraer_fecha(texto):
    """Extrae la primera fecha encontrada en el texto (formato DD/MM/AA o DD/MM/AAAA)."""
    # Patrón: dd/mm/yy o dd/mm/yyyy
    match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', texto)
    if match:
        dia, mes, anio = match.groups()
        # Si el año es de 2 dígitos, asumir 2000+
        if len(anio) == 2:
            anio = "20" + anio
        try:
            fecha = datetime(int(anio), int(mes), int(dia))
            return fecha.strftime("%Y-%m-%d")
        except ValueError:
            return match.group(0)  # Devolver tal cual si no es válida

    # Buscar patrón con texto: "6 de abril", "15 de mayo", etc.
    meses_texto = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    match2 = re.search(r'(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)', texto.lower())
    if match2:
        dia = int(match2.group(1))
        mes = meses_texto[match2.group(2)]
        try:
            fecha = datetime(2026, mes, dia)
            return fecha.strftime("%Y-%m-%d")
        except ValueError:
            pass

    return ""


def clasificar_tipo(texto):
    """
    Clasifica el tipo de evento basándose en el contenido del texto.
    Tipos posibles:
    - Capacitación
    - Incidencia Técnica - Software
    - Incidencia Técnica - Hardware
    - Incidencia Técnica - Red/Conectividad
    - Seguimiento
    - Visita Técnica
    - Instalación
    - Informativo
    """
    texto_lower = texto.lower()

    # Capacitación (NO es incidencia)
    palabras_capacitacion = ['capacitación', 'capacitacion', 'capacitaci']
    es_capacitacion = any(p in texto_lower for p in palabras_capacitacion)

    # Si SOLO habla de capacitación sin mencionar problemas técnicos
    palabras_incidencia = [
        'falla', 'problema', 'no funciona', 'no responde', 'pantalla negra',
        'no arrancan', 'no carga', 'latencia', 'lento', 'recalentamiento',
        'no entra', 'no abre', 'no dej', 'sin sistema', 'pila del bios',
        'login', 'grub', 'error'
    ]
    tiene_incidencia = any(p in texto_lower for p in palabras_incidencia)

    palabras_red = [
        'internet', 'red', 'conectividad', 'cableado', 'wifi',
        'conexión', 'conexion', 'redpe', 'otic'
    ]
    tiene_red = any(p in texto_lower for p in palabras_red)

    palabras_hardware = [
        'hardware', 'pila', 'bios', 'pantalla negra', 'recalentamiento',
        'sonido', 'hdmi', 'proyección', 'proyeccion', 'no arrancan',
        'corriente', 'setup'
    ]
    tiene_hardware = any(p in texto_lower for p in palabras_hardware)

    palabras_software = [
        'sistema operativo', 'inkscape', 'programa', 'office', 'publisher',
        'access', 'italc', 'dfd', 'iso', 'actualiz', 'instalar programa',
        'apps', 'software', 'computador virtual', 'compuvirtual'
    ]
    tiene_software = any(p in texto_lower for p in palabras_software)

    palabras_instalacion = [
        'se instaló', 'se instalo', 'instalación completa', 'instalacion completa',
        'no se instaló', 'no se instalo', 'visita de instalación', 'no se pudo instalar'
    ]
    tiene_instalacion = any(p in texto_lower for p in palabras_instalacion)

    palabras_visita = ['visita tecnica', 'visita técnica']
    tiene_visita = any(p in texto_lower for p in palabras_visita)

    # Priorización de clasificación
    if es_capacitacion and not tiene_incidencia and not tiene_red:
        return "Capacitación"

    if tiene_instalacion:
        return "Instalación"

    if tiene_visita and (tiene_incidencia or tiene_red or tiene_hardware):
        return "Visita Técnica"

    if tiene_red and (tiene_incidencia or 'solicitan' in texto_lower or 'arreglo' in texto_lower):
        return "Incidencia Técnica - Red/Conectividad"

    if tiene_hardware:
        return "Incidencia Técnica - Hardware"

    if tiene_software and tiene_incidencia:
        return "Incidencia Técnica - Software"

    if tiene_software and not tiene_incidencia:
        return "Gestión de Software"

    if tiene_incidencia:
        return "Incidencia Técnica - General"

    if tiene_red:
        return "Incidencia Técnica - Red/Conectividad"

    if es_capacitacion:
        return "Capacitación"

    # Revisar si es un seguimiento sin novedad
    if 'no reportan' in texto_lower and 'novedad' in texto_lower:
        return "Seguimiento sin novedad"

    if 'seguimiento' in texto_lower:
        return "Seguimiento"

    if 'pendiente' in texto_lower:
        return "Pendiente"

    return "Informativo"


def detectar_responsable(texto):
    """Detecta el responsable mencionado en el texto del evento."""
    texto_lower = texto.lower().strip()

    # Buscar al final del texto o en el contenido
    if texto_lower.endswith('oit') or 'oit' in texto_lower.split('.')[-1].lower():
        return "OIT"
    if texto_lower.endswith('qinaya') or 'qinaya' in texto_lower.split('.')[-1].lower():
        return "Qinaya"
    if 'oit y qinaya' in texto_lower or 'qinaya y oit' in texto_lower:
        return "OIT y Qinaya"
    if ' oit' in texto_lower or 'oit ' in texto_lower:
        return "OIT"
    if 'qinaya' in texto_lower:
        return "Qinaya"

    return ""


def separar_eventos(texto_bot):
    """
    Separa el texto de la columna Bot en eventos individuales.
    Cada línea separada por \n que contenga contenido sustancial es un evento.
    También maneja líneas de continuación (sin fecha propia) uniéndolas
    al evento anterior.
    """
    if not texto_bot:
        return [{"texto": "", "fecha": "", "tipo": "Sin registro"}]

    lineas = [l.strip() for l in texto_bot.split('\n') if l.strip()]

    if not lineas:
        return [{"texto": texto_bot.strip(), "fecha": extraer_fecha(texto_bot), "tipo": clasificar_tipo(texto_bot)}]

    eventos = []
    evento_actual = None

    for linea in lineas:
        # Determinar si es un nuevo evento (empieza con fecha, "Seguimiento", "Visita", "Solicitan", "Pendiente", "Novedad", "Capacitación", o tiene fecha al inicio)
        es_nuevo_evento = bool(re.match(
            r'^(Seguimiento|Visita|Solicitan|Pendiente|Novedad|Capacitaci|Se realiz|Se instal|Se envi|Coloco|Se hizo|\d{1,2}/\d{1,2}/\d{2,4})',
            linea, re.IGNORECASE
        ))

        if es_nuevo_evento:
            # Guardar evento anterior si existe
            if evento_actual:
                eventos.append(evento_actual)
            evento_actual = linea
        else:
            # Es continuación del evento anterior
            if evento_actual:
                evento_actual += " " + linea
            else:
                evento_actual = linea

    # Agregar último evento
    if evento_actual:
        eventos.append(evento_actual)

    resultado = []
    for ev in eventos:
        fecha = extraer_fecha(ev)
        tipo = clasificar_tipo(ev)
        responsable = detectar_responsable(ev)
        resultado.append({
            "texto": ev,
            "fecha": fecha,
            "tipo": tipo,
            "responsable": responsable
        })

    return resultado


def main():
    print("[INFO] Leyendo archivo original 'Reporte de Incidencias.xlsx'...")
    wb = openpyxl.load_workbook('Reporte de Incidencias.xlsx')
    ws = wb.active

    # Crear nuevo workbook
    wb_nuevo = openpyxl.Workbook()
    ws_nuevo = wb_nuevo.active
    ws_nuevo.title = "Incidencias Normalizado"

    ws_cap = wb_nuevo.create_sheet("Capacitaciones")
    ws_sin_novedad = wb_nuevo.create_sheet("Sin Novedad")

    # ==========================================
    # ENCABEZADOS
    # ==========================================
    headers = [
        "ID",
        "Código DANE",
        "Colegio",
        "Docentes",
        "WhatsApp Creado",
        "Fase",
        "Fecha del Evento",
        "Descripción del Evento",
        "Clasificación del Evento",
        "Responsable",
        "Tipo Incidencia Original",
        "Estado General"
    ]

    # Estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        # Para hoja de incidencias
        cell = ws_nuevo.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Para hoja de capacitaciones
        cell_cap = ws_cap.cell(row=1, column=col, value=header)
        cell_cap.font = header_font
        cell_cap.fill = header_fill
        cell_cap.alignment = header_alignment
        cell_cap.border = thin_border
        
        # Para hoja sin novedad
        cell_sn = ws_sin_novedad.cell(row=1, column=col, value=header)
        cell_sn.font = header_font
        cell_sn.fill = header_fill
        cell_sn.alignment = header_alignment
        cell_sn.border = thin_border

    # ==========================================
    # PROCESAR DATOS
    # ==========================================
    fila_destino = 2
    fila_destino_cap = 2
    fila_destino_sn = 2
    id_evento = 1
    total_colegios = 0
    total_eventos = 0

    for fila in range(2, ws.max_row + 1):
        colegio = ws.cell(fila, 1).value
        if not colegio:
            break

        total_colegios += 1
        docentes = ws.cell(fila, 2).value or 0
        whatsapp = ws.cell(fila, 3).value or False
        fase = ws.cell(fila, 4).value or ""
        bot = ws.cell(fila, 5).value or ""
        tipo_incidencia_orig = ws.cell(fila, 6).value or ""
        # responsable_orig = ws.cell(fila, 7).value or ""  # Not used directly
        estado = ws.cell(fila, 8).value or ""

        codigo_dane = buscar_dane(colegio)
        eventos = separar_eventos(bot)

        for evento in eventos:
            total_eventos += 1
            
            # Determinar a qué hoja va
            es_capacitacion = ("Capacitación" in evento["tipo"])
            es_sin_novedad = ("sin novedad" in evento["tipo"].lower())
            
            if es_capacitacion:
                ws_actual = ws_cap
                fila_actual = fila_destino_cap
            elif es_sin_novedad:
                ws_actual = ws_sin_novedad
                fila_actual = fila_destino_sn
            else:
                ws_actual = ws_nuevo
                fila_actual = fila_destino

            ws_actual.cell(row=fila_actual, column=1, value=id_evento)
            ws_actual.cell(row=fila_actual, column=2, value=codigo_dane)
            ws_actual.cell(row=fila_actual, column=3, value=colegio)
            ws_actual.cell(row=fila_actual, column=4, value=int(docentes) if docentes else 0)
            ws_actual.cell(row=fila_actual, column=5, value="Sí" if whatsapp else "No")
            ws_actual.cell(row=fila_actual, column=6, value=int(fase) if fase else "")
            ws_actual.cell(row=fila_actual, column=7, value=evento["fecha"])
            ws_actual.cell(row=fila_actual, column=8, value=evento["texto"])
            ws_actual.cell(row=fila_actual, column=9, value=evento["tipo"])
            ws_actual.cell(row=fila_actual, column=10, value=evento.get("responsable", ""))
            ws_actual.cell(row=fila_actual, column=11, value=tipo_incidencia_orig)
            ws_actual.cell(row=fila_actual, column=12, value=estado)

            # Aplicar bordes y formato
            for col in range(1, len(headers) + 1):
                cell = ws_actual.cell(row=fila_actual, column=col)
                cell.border = thin_border
                cell.font = Font(name='Calibri', size=10)
                if col in [1, 2, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif col == 8:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='top')

            # Color de fondo para clasificación
            tipo_cell = ws_actual.cell(row=fila_actual, column=9)
            tipo_val = evento["tipo"]
            if "Capacitación" in tipo_val:
                tipo_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            elif "Red" in tipo_val:
                tipo_cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            elif "Hardware" in tipo_val:
                tipo_cell.fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
            elif "Software" in tipo_val:
                tipo_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            elif "sin novedad" in tipo_val.lower():
                tipo_cell.fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')

            if es_capacitacion:
                fila_destino_cap += 1
            elif es_sin_novedad:
                fila_destino_sn += 1
            else:
                fila_destino += 1
                
            id_evento += 1

    # ==========================================
    # AJUSTAR ANCHOS DE COLUMNA
    # ==========================================
    anchos = {
        1: 6,    # ID
        2: 16,   # Código DANE
        3: 35,   # Colegio
        4: 10,   # Docentes
        5: 12,   # WhatsApp
        6: 8,    # Fase
        7: 14,   # Fecha
        8: 80,   # Descripción
        9: 32,   # Clasificación
        10: 14,  # Responsable
        11: 35,  # Tipo Original
        12: 14,  # Estado
    }
    for col, ancho in anchos.items():
        ws_nuevo.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
        ws_cap.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
        ws_sin_novedad.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    # Congelar primera fila
    ws_nuevo.freeze_panes = 'A2'
    ws_cap.freeze_panes = 'A2'
    ws_sin_novedad.freeze_panes = 'A2'

    # Filtros automáticos
    if fila_destino > 2:
        ws_nuevo.auto_filter.ref = f"A1:L{fila_destino - 1}"
    if fila_destino_cap > 2:
        ws_cap.auto_filter.ref = f"A1:L{fila_destino_cap - 1}"
    if fila_destino_sn > 2:
        ws_sin_novedad.auto_filter.ref = f"A1:L{fila_destino_sn - 1}"

    # ==========================================
    # HOJA RESUMEN
    # ==========================================
    ws_resumen = wb_nuevo.create_sheet("Resumen")
    ws_resumen.cell(1, 1, "Resumen de Transformación").font = Font(bold=True, size=14, color='1F4E79')
    ws_resumen.cell(3, 1, "Total Colegios:")
    ws_resumen.cell(3, 2, total_colegios)
    ws_resumen.cell(4, 1, "Total Eventos Generados:")
    ws_resumen.cell(4, 2, total_eventos)
    ws_resumen.cell(5, 1, "Fecha de Generación:")
    ws_resumen.cell(5, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))

    # Contar por tipo
    tipos_count = {}
    for fila in range(2, fila_destino):
        tipo = ws_nuevo.cell(fila, 9).value
        tipos_count[tipo] = tipos_count.get(tipo, 0) + 1

    ws_resumen.cell(7, 1, "Distribución por Tipo de Evento:").font = Font(bold=True, size=11)
    fila_res = 8
    for tipo, count in sorted(tipos_count.items(), key=lambda x: -x[1]):
        ws_resumen.cell(fila_res, 1, tipo)
        ws_resumen.cell(fila_res, 2, count)
        fila_res += 1

    ws_resumen.column_dimensions['A'].width = 40
    ws_resumen.column_dimensions['B'].width = 15

    # ==========================================
    # GUARDAR
    # ==========================================
    output_file = 'Reporte_Incidencias_Normalizado.xlsx'
    wb_nuevo.save(output_file)

    print(f"\n[OK] Transformacion completada exitosamente!")
    print(f"   - {total_colegios} colegios procesados")
    print(f"   - {total_eventos} eventos individuales generados")
    print(f"   - Archivo guardado: {output_file}")
    print(f"\n[RESUMEN] Distribucion por tipo:")
    for tipo, count in sorted(tipos_count.items(), key=lambda x: -x[1]):
        print(f"   * {tipo}: {count}")


if __name__ == "__main__":
    main()
