import json
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# Cargar datos
with open('data_indicadores.json', encoding='utf-8') as f:
    data = json.load(f)

resumen = data['resumen']
detalle = data['detalle_colegios']

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
#  HOJA 1 – RESUMEN EJECUTIVO
# ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Resumen Ejecutivo"

# Colores
AZUL_OSC   = "0F172A"
AZUL_MED   = "1E3A5F"
AZUL_ACT   = "00D2FF"
BLANCO     = "FFFFFF"
GRIS_CLAR  = "F1F5F9"
VERDE      = "10B981"
AMARILLO   = "F59E0B"
ROJO       = "EF4444"
GRIS_TEXT  = "64748B"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def border_med():
    s = Side(style='medium', color="00D2FF")
    return Border(left=s, right=s, top=s, bottom=s)

# Título principal - fila 1 (merge A1:E1)
ws1.merge_cells("A1:E1")
c = ws1["A1"]
c.value = "INDICADOR Q-07 — AUTONOMÍA OPERATIVA"
c.font = Font(name="Calibri", bold=True, size=18, color=BLANCO)
c.fill = fill(AZUL_OSC)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 38

# Subtítulo fila 2
ws1.merge_cells("A2:E2")
c = ws1["A2"]
c.value = "Período: 13 Abril 2026 – 15 Junio 2026  |  Colegios con instalación activa del sistema Qinaya"
c.font = Font(name="Calibri", italic=True, size=11, color="94A3B8")
c.fill = fill(AZUL_OSC)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 22

# Fila vacía
ws1.row_dimensions[3].height = 10

# ENCABEZADO KPIs (fila 4)
kpi_headers = ["📊 KPI", "Valor", "Descripción"]
for col_i, h in enumerate(kpi_headers, 1):
    c = ws1.cell(row=4, column=col_i, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
    c.fill = fill(AZUL_MED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
ws1.row_dimensions[4].height = 24

# Datos KPI
kpis = [
    ("🏫 Colegios Evaluados",           f"{resumen['total_colegios']} colegios",       "Con instalación activa del sistema Qinaya"),
    ("📅 Días Hábiles del Periodo",     f"{resumen['dias_habiles_periodo']} días",      "Sin contar sábados, domingos ni festivos"),
    ("🖥️ Sesiones Estimadas por Col.",  f"{resumen['sesiones_mes_estimadas']} sesiones","Aprox. 3.5 sesiones/día × 44 días hábiles"),
    ("🔧 Soportes Promedio por Col.",   f"{resumen['promedio_soportes_colegio_mes']}", "Total soportes ÷ 30 colegios instalados"),
    ("✅ Indicador Global Q-07",         f"{resumen['promedio_autonomia_global']}%",    "% sesiones realizadas sin requerir soporte"),
]

for r_i, (kpi, val, desc) in enumerate(kpis, 5):
    ws1.cell(row=r_i, column=1, value=kpi).font = Font(name="Calibri", bold=True, size=11, color=AZUL_OSC)
    ws1.cell(row=r_i, column=1).fill = fill(GRIS_CLAR)
    ws1.cell(row=r_i, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.cell(row=r_i, column=1).border = border_thin()

    val_cell = ws1.cell(row=r_i, column=2, value=val)
    val_cell.font = Font(name="Calibri", bold=True, size=13, color="00D2FF")
    val_cell.fill = fill(AZUL_OSC)
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.border = border_thin()

    desc_cell = ws1.cell(row=r_i, column=3, value=desc)
    desc_cell.font = Font(name="Calibri", size=10, color=GRIS_TEXT)
    desc_cell.fill = fill(GRIS_CLAR)
    desc_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    desc_cell.border = border_thin()
    ws1.row_dimensions[r_i].height = 26

# Fórmula explicada (fila 11)
ws1.row_dimensions[11].height = 12
ws1.merge_cells("A12:E12")
c = ws1["A12"]
c.value = "FÓRMULA:  Q-07 = (Sesiones sin soporte / Total sesiones estimadas) × 100"
c.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
c.fill = fill(AZUL_MED)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[12].height = 22

# Nota sobre semáforo
ws1.row_dimensions[13].height = 8
ws1.merge_cells("A14:E14")
c = ws1["A14"]
c.value = "SEMÁFORO:  ✅ ≥ 95%  Óptimo     ⚠️ 90–94%  Aceptable     🔴 < 90%  Requiere atención"
c.font = Font(name="Calibri", size=10, color=AZUL_OSC)
c.fill = fill("FFF9C4")
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[14].height = 20

# Ajustar anchos col hoja 1
ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 48
ws1.column_dimensions["D"].width = 5
ws1.column_dimensions["E"].width = 5

# ─────────────────────────────────────────────
#  HOJA 2 – DETALLE POR COLEGIO
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Detalle por Colegio")

# Título
ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "Q-07 — DETALLE POR INSTITUCIÓN EDUCATIVA"
c.font = Font(name="Calibri", bold=True, size=15, color=BLANCO)
c.fill = fill(AZUL_OSC)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

# Subtítulo
ws2.merge_cells("A2:F2")
c = ws2["A2"]
c.value = "Período: 13 Abril – 15 Junio 2026   |   Solo colegios con sistema Qinaya instalado"
c.font = Font(name="Calibri", italic=True, size=10, color="94A3B8")
c.fill = fill(AZUL_OSC)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 18

ws2.row_dimensions[3].height = 8

# Encabezados tabla
headers = [
    "#",
    "Institución Educativa",
    "Sesiones Estimadas",
    "Soportes Requeridos",
    "Sesiones Exitosas",
    "Indicador Q-07 (%)",
]
for col_i, h in enumerate(headers, 1):
    c = ws2.cell(row=4, column=col_i, value=h)
    c.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
    c.fill = fill("1E3A5F")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_thin()
ws2.row_dimensions[4].height = 32

# Datos
for r_i, item in enumerate(detalle, 5):
    idx = r_i - 4
    bg = GRIS_CLAR if idx % 2 == 0 else BLANCO

    # Número
    c = ws2.cell(row=r_i, column=1, value=idx)
    c.font = Font(name="Calibri", bold=True, size=10, color=GRIS_TEXT)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()

    # Colegio
    c = ws2.cell(row=r_i, column=2, value=item["colegio"])
    c.font = Font(name="Calibri", bold=True, size=10, color=AZUL_OSC)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = border_thin()

    # Sesiones estimadas
    c = ws2.cell(row=r_i, column=3, value=item["total_sesiones_estimadas"])
    c.font = Font(name="Calibri", size=10, color=AZUL_OSC)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
    c.number_format = "#,##0"

    # Soportes
    sop = item["soportes_requeridos"]
    c = ws2.cell(row=r_i, column=4, value=sop)
    if sop >= 5:
        c.font = Font(name="Calibri", bold=True, size=10, color=ROJO)
        c.fill = fill("FEE2E2")
    elif sop >= 3:
        c.font = Font(name="Calibri", bold=True, size=10, color=AMARILLO)
        c.fill = fill("FEF9C3")
    elif sop >= 1:
        c.font = Font(name="Calibri", size=10, color="D97706")
        c.fill = fill(bg)
    else:
        c.font = Font(name="Calibri", size=10, color=VERDE)
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()

    # Sesiones exitosas
    c = ws2.cell(row=r_i, column=5, value=item["sesiones_sin_soporte"])
    c.font = Font(name="Calibri", size=10, color=AZUL_OSC)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
    c.number_format = "#,##0"

    # Indicador Q-07
    q = item["indicador_q07"]
    c = ws2.cell(row=r_i, column=6, value=q / 100)
    if q >= 95:
        c.font = Font(name="Calibri", bold=True, size=11, color=VERDE)
        c.fill = fill("D1FAE5")
    elif q >= 90:
        c.font = Font(name="Calibri", bold=True, size=11, color=AMARILLO)
        c.fill = fill("FEF9C3")
    else:
        c.font = Font(name="Calibri", bold=True, size=11, color=ROJO)
        c.fill = fill("FEE2E2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
    c.number_format = "0.00%"

    ws2.row_dimensions[r_i].height = 22

# Fila de totales / promedio
last_row = 4 + len(detalle) + 1
ws2.merge_cells(f"A{last_row}:B{last_row}")
c = ws2.cell(row=last_row, column=1, value="PROMEDIO GLOBAL Q-07")
c.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
c.fill = fill(AZUL_OSC)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = border_thin()

ws2.cell(row=last_row, column=3).value = resumen["sesiones_mes_estimadas"]
ws2.cell(row=last_row, column=3).font = Font(name="Calibri", bold=True, color=BLANCO)
ws2.cell(row=last_row, column=3).fill = fill(AZUL_OSC)
ws2.cell(row=last_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
ws2.cell(row=last_row, column=3).border = border_thin()
ws2.cell(row=last_row, column=3).number_format = "#,##0"

ws2.cell(row=last_row, column=4).value = resumen["promedio_soportes_colegio_mes"]
ws2.cell(row=last_row, column=4).font = Font(name="Calibri", bold=True, color="00D2FF")
ws2.cell(row=last_row, column=4).fill = fill(AZUL_OSC)
ws2.cell(row=last_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
ws2.cell(row=last_row, column=4).border = border_thin()
ws2.cell(row=last_row, column=4).number_format = "0.00"

ws2.cell(row=last_row, column=5).fill = fill(AZUL_OSC)
ws2.cell(row=last_row, column=5).border = border_thin()

c_prom = ws2.cell(row=last_row, column=6, value=resumen["promedio_autonomia_global"] / 100)
c_prom.font = Font(name="Calibri", bold=True, size=13, color="00D2FF")
c_prom.fill = fill(AZUL_OSC)
c_prom.alignment = Alignment(horizontal="center", vertical="center")
c_prom.border = border_thin()
c_prom.number_format = "0.00%"
ws2.row_dimensions[last_row].height = 28

# Anchos de columnas hoja 2
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 22
ws2.column_dimensions["F"].width = 22

# Freeze encabezados
ws2.freeze_panes = "A5"

# Guardar
wb.save("Q07_Autonomia_Operativa.xlsx")
print("Archivo Excel generado: Q07_Autonomia_Operativa.xlsx")
